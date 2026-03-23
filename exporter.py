from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import datetime
import os

def export_to_excel(companies, category_name, export_dir="exports"):
    """Génère un fichier .xlsx à partir de la liste des entreprises."""
    if not os.path.exists(export_dir):
        os.makedirs(export_dir)
        
    wb = Workbook()
    ws = wb.active
    ws.title = "Entreprises"
    
    # En-têtes
    headers = ["Nom de l'entreprise", "Téléphone", "Site Web", "Catégorie", "Coordonnées GPS", "Lien Annuaire"]
    ws.append(headers)
    
    # Style pour les en-têtes
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Données
    for company in companies:
        ws.append([
            company.get("name"),
            company.get("phone"),
            company.get("website"),
            category_name,
            company.get("coords", "N/A"),
            company.get("detail_url")
        ])
    
    # Ajuster la largeur des colonnes
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 50)
        
    # Nom du fichier
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = "".join([c for c in category_name if c.isalnum() or c in (' ', '_')]).rstrip().replace(' ', '_')
    filename = f"entreprises_{safe_name}_{timestamp}.xlsx"
    filepath = os.path.join(export_dir, filename)
    
    wb.save(filepath)
    return filename, filepath
